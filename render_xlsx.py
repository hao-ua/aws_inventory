import os
import openpyxl
from openpyxl.cell import get_column_letter
from openpyxl.styles import Alignment, Font, Style


class Render(object):
    def __init__(self, output_folder, categories, pages):
        self.categories = categories
        self.items = {}
        self.Name = 'XLSX'
        self.HeaderRows = {}
        self.HeaderRowWidths = {}
        self.HeaderKeys = {}
        self.START_ROW = 1
        self.START_COLUMN = 1
        self.currentRow = self.START_ROW
        self.printAddFunctions = {}
        self.pages = pages
        for page in self.pages:
            self.printAddFunctions[page] = None

        self.output_folder = output_folder

    def config_page(self, page, header_rows, header_row_widths, header_keys, data):
        self.HeaderRows[page] = header_rows
        self.HeaderRowWidths[page] = header_row_widths
        self.HeaderKeys[page] = header_keys
        self.items[page] = data

    def print_row(self, page, rows, data, bold=False):
        items = [str(data[key]) for key in rows]
        if len(items) == 1 and len(items) != len(self.HeaderRows[page.title]):
            page.merge_cells(start_row=self.currentRow, start_column=self.START_COLUMN, end_row=self.currentRow,
                             end_column=len(self.HeaderRows[page.title]))

        for index, val in enumerate(items, start=self.START_COLUMN):
            cell = page.cell(row=self.currentRow, column=index)
            cell.value = val
            if bold:
                cell.style = Style(font=Font(bold=bold), alignment=Alignment(horizontal='center', vertical='center',
                                                                             wrap_text=True))
            else:
                cell.style = Style(font=Font(bold=bold), alignment=Alignment(wrap_text=True))

        self.currentRow += 1

    def print_category_items(self, category, page):
        data = {}
        for key, val in map(lambda i, j: (i, j), self.HeaderKeys[page.title], self.HeaderRows[page.title]):
            data[key] = val

        self.print_row(page, self.HeaderKeys[page.title], data, True)
        key_list = sorted(self.items[page.title][category].keys())
        for region_name in key_list:
            if len(self.items[page.title][category][region_name]) != 0:
                self.print_row(page, ['region'], {'region': region_name.upper()}, True)
                item_list = sorted(self.items[page.title][category][region_name], key=lambda x: x['name'].upper())
                for item in item_list:
                    self.print_row(page, self.HeaderKeys[page.title], item)

    def generate_page(self, page):
        for category in self.categories:
            self.print_row(page, ['category'], {'category': category}, True)
            self.print_category_items(category, page)
            self.currentRow += 1

    def render(self):
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = self.pages[0]
        for page in self.pages[1:]:
            ws = wb.create_sheet()
            ws.title = page

        for page in self.pages:
            ws = wb.get_sheet_by_name(page)
            self.currentRow = self.START_ROW
            self.generate_page(ws)
            for i, val in enumerate(self.HeaderRowWidths[page]):
                ws.column_dimensions[get_column_letter(i+1)].width = int(val)*11

        wb.save(''.join([self.output_folder, os.sep, 'inventory.xlsx']))
